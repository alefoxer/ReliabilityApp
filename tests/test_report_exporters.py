import io
from datetime import datetime
from zipfile import ZipFile

import pytest
from PIL import Image

from app.core.rbd_models import ReportData
from app.reports.report_exporters import (
    _add_xlsx_image_with_reserved_space,
    _formula_png_bytes,
    _formula_render_blocks,
    _normalize_image_dimensions,
    export_docx,
    export_html,
    export_txt,
    export_xlsx,
)


def _sample_report() -> ReportData:
    return ReportData(
        title="Отчёт по расчёту",
        subtitle="Инженерный отчёт",
        created_at=datetime(2026, 4, 21, 12, 0),
        inputs={"lambda": 0.001, "t": 1000, "Способ расчёта": "Аналитический расчёт"},
        results={"P": 0.9, "T0": 1000},
        method_name="F1.1",
        methodology="Нормативная формула проекта",
        calculation_method="Аналитический расчёт",
        formula_text="P(t)=e^{-lambda t}",
        formula_latex=r"\[ P(t)=e^{-\lambda t} \]",
        tables={"Данные графика": [(0, 1.0), (100, 0.95), (1000, 0.9)]},
        threshold_metric="P",
        threshold_value=0.8,
        threshold_passed=True,
        threshold_conclusion="Система соответствует заданному порогу.",
        final_conclusion="Расчёт подтверждает выполнение критерия надёжности.",
        warnings=["Тестовое предупреждение"],
        limitations=["Тестовое ограничение"],
        metadata={
            "nomenclature": {
                "purpose_label": "Конкретное назначение",
                "usage_mode_label": "Непрерывное длительное применение",
                "recovery_mode_label": "Восстанавливаемое обслуживаемое изделие",
                "requires_tto": True,
                "tto": "2.5",
                "recommended_metrics_text": "P(t), T0, Kг, Tто",
            }
        },
    )


def _sample_threshold_tv_report() -> ReportData:
    report = _sample_report()
    report.results = {"Tv": 8.0, "Kg": 0.92}
    report.threshold_metric = "Tv"
    report.threshold_value = 10.0
    report.threshold_passed = True
    report.threshold_conclusion = "Система соответствует заданному порогу: Tv=8.000000, порог=10.000000."
    report.final_conclusion = report.threshold_conclusion
    return report


def _write_png(path):
    path.write_bytes(
        bytes.fromhex(
            "89504E470D0A1A0A"
            "0000000D49484452000000010000000108060000001F15C489"
            "0000000D49444154789C63F8FFFFFF7F0009FB03FD2A86E38A"
            "0000000049454E44AE426082"
        )
    )
    return str(path)


def _sample_scheme_images(tmp_path):
    main = _write_png(tmp_path / "main.png")
    child = _write_png(tmp_path / "child.png")
    nested = _write_png(tmp_path / "nested.png")
    return [
        {"title": "Схема системы", "path": main, "level": 0, "block_name": ""},
        {"title": "Подсхема блока Блок 1", "path": child, "level": 1, "block_name": "Блок 1"},
        {"title": "Подсхема блока Блок 1.1", "path": nested, "level": 2, "block_name": "Блок 1.1"},
    ]


def test_export_html_contains_report_sections(tmp_path):
    target = export_html(tmp_path / "report.html", _sample_report())
    content = target.read_text(encoding="utf-8")

    assert "<!DOCTYPE html>" in content
    assert "Надёжность технических средств" in content
    assert "Итоговый вывод" in content
    assert "Исходные данные" in content
    assert "Числовые результаты" in content
    assert "Формулы" in content
    assert "Номенклатура показателей" in content
    assert "Система соответствует заданному порогу." in content
    assert "Способ расчёта" in content


def test_export_html_uses_svg_formula_rendering(tmp_path):
    target = export_html(tmp_path / "report_svg.html", _sample_report())
    content = target.read_text(encoding="utf-8")

    assert "data:image/svg+xml;base64," in content


def test_export_html_includes_subscheme_sections(tmp_path):
    report = _sample_report()
    report.scheme_image_path = _write_png(tmp_path / "legacy_main.png")
    report.scheme_images = _sample_scheme_images(tmp_path)

    target = export_html(tmp_path / "report_subschemes.html", report)
    content = target.read_text(encoding="utf-8")

    assert "Схема системы" in content
    assert "Подсхема блока Блок 1" in content
    assert "Подсхема блока Блок 1.1" in content


def test_export_html_groups_main_scheme_and_subschemes_in_one_block(tmp_path):
    report = _sample_report()
    report.scheme_image_path = _write_png(tmp_path / "legacy_main_grouped.png")
    report.scheme_images = _sample_scheme_images(tmp_path)

    target = export_html(tmp_path / "report_subschemes_grouped.html", report)
    content = target.read_text(encoding="utf-8")

    assert content.count("<h2>Схема системы</h2>") == 1
    assert "<h3>Подсхема блока Блок 1</h3>" in content
    assert "<h3>Подсхема блока Блок 1.1</h3>" in content


def test_export_html_uses_selected_non_p_threshold_metric(tmp_path):
    target = export_html(tmp_path / "report_tv.html", _sample_threshold_tv_report())
    content = target.read_text(encoding="utf-8")

    assert "Tv=8.000000" in content
    assert "порог=10.000000" in content
    assert "Система соответствует заданному порогу" in content


def test_export_html_does_not_show_internal_auto_id_as_method_or_formula(tmp_path):
    report = _sample_report()
    report.method_name = "Схема: AUTO.COMPOSITION - Автоматическая общая формула по схеме"
    report.formula_text = "Pсист(t)=B1\nAUTO.COMPOSITION"
    report.formula_latex = ""
    target = export_html(tmp_path / "report.html", report)
    content = target.read_text(encoding="utf-8")

    assert "AUTO.COMPOSITION" not in content
    assert "Pсист(t)=B1" in content
    assert "Формула по структуре схемы" in content


def test_export_docx_contains_core_sections(tmp_path):
    pytest.importorskip("docx")
    target = export_docx(tmp_path / "report.docx", _sample_report())

    with ZipFile(target) as archive:
        names = set(archive.namelist())
        document_xml = archive.read("word/document.xml").decode("utf-8")

    assert "word/document.xml" in names
    assert "Краткая сводка" in document_xml
    assert "Результаты расчёта" in document_xml
    assert "Формулы" in document_xml
    assert "Номенклатура показателей" in document_xml
    assert "Заключение" in document_xml
    assert any(name.startswith("word/media/") for name in names)
    assert "\\[" not in document_xml
    assert target.stat().st_size > 0


def test_export_docx_includes_subscheme_titles(tmp_path):
    pytest.importorskip("docx")
    report = _sample_report()
    report.scheme_images = _sample_scheme_images(tmp_path)
    report.scheme_image_path = report.scheme_images[0]["path"]
    target = export_docx(tmp_path / "report_subschemes.docx", report)

    with ZipFile(target) as archive:
        document_xml = archive.read("word/document.xml").decode("utf-8")

    assert "Подсхема блока Блок 1" in document_xml
    assert "Подсхема блока Блок 1.1" in document_xml


def test_export_docx_formula_fallback_uses_plain_text(tmp_path, monkeypatch):
    pytest.importorskip("docx")
    monkeypatch.setattr("app.reports.report_exporters._formula_png_bytes", lambda block: b"")
    target = export_docx(tmp_path / "report_fallback.docx", _sample_report())

    with ZipFile(target) as archive:
        document_xml = archive.read("word/document.xml").decode("utf-8")

    assert "P(t)=e^-lambda t" in document_xml
    assert "\\[" not in document_xml


def test_export_xlsx_contains_professional_sheets(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    target = export_xlsx(tmp_path / "report.xlsx", _sample_report())
    workbook = openpyxl.load_workbook(target)

    assert workbook.sheetnames[0] == "Титульный лист"
    assert {
        "Сводка",
        "Паспорт отчёта",
        "Исходные данные",
        "Результаты",
        "Формулы",
        "Методика",
        "Номенклатура",
        "Данные графиков",
        "Графики",
        "Схема",
        "Проверки",
        "Заключение",
    }.issubset(set(workbook.sheetnames))
    assert workbook["Сводка"]["A3"].value == "Параметр"
    assert workbook["Номенклатура"]["A2"].value == "Назначение"
    assert workbook["Проверки"]["A2"].value == "Пороговый показатель"
    with ZipFile(target) as archive:
        names = set(archive.namelist())
        formula_image_name = next(name for name in names if name.startswith("xl/media/"))
        formula_image_data = archive.read(formula_image_name)
    assert any(name.startswith("xl/media/") for name in names)
    assert not any(name.startswith("xl/charts/") for name in names)
    image = Image.open(io.BytesIO(formula_image_data))
    assert image.mode == "RGB"
    assert target.stat().st_size > 0


def test_export_xlsx_title_sheet_has_navigation_links(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    target = export_xlsx(tmp_path / "report_navigation.xlsx", _sample_report())
    workbook = openpyxl.load_workbook(target)
    title_sheet = workbook["Титульный лист"]

    links = [cell.hyperlink.target for row in title_sheet.iter_rows() for cell in row if cell.hyperlink]
    assert "#'Сводка'!A1" in links
    assert "#'Формулы'!A1" in links
    assert workbook["Сводка"]["H1"].hyperlink.target == "#'Титульный лист'!A1"


def test_xlsx_image_helper_reserves_rows_and_limits_upscale(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XlsxImage

    path = tmp_path / "small.png"
    Image.new("RGB", (40, 20), "white").save(path)
    workbook = Workbook()
    sheet = workbook.active

    next_row = _add_xlsx_image_with_reserved_space(
        sheet,
        str(path),
        2,
        2,
        XlsxImage,
        max_width_px=400,
        max_height_px=200,
        max_upscale=1.1,
    )

    assert next_row > 2
    assert len(sheet._images) == 1
    assert sheet._images[0].width <= 44
    assert sheet._images[0].height <= 22
    assert sheet.row_dimensions[2].height is not None


def test_formula_report_rendering_keeps_russian_text_out_of_math_mode():
    text_block = {
        "label": "Описание",
        "latex": "Структурная формула\nВероятность безотказной работы\nПоследовательное соединение",
        "plain": "Структурная формула\nВероятность безотказной работы\nПоследовательное соединение",
    }
    formula_block = {
        "label": "P",
        "latex": r"P_{\text{сист}}(t)=P_{B1}\cdot P_{B2}",
        "plain": "Pсист(t)=P_B1 * P_B2",
    }

    assert _formula_png_bytes(text_block) == b""
    assert _formula_png_bytes(formula_block)


def test_formula_package_aggregate_text_is_not_rendered_as_one_image():
    from app.core.rbd_models import FormulaItem, FormulaPackage

    package = FormulaPackage(
        formula_mode="structural_fallback",
        is_normative=False,
        method_code=None,
        title="",
        source_label="",
        source_details="",
        applicability="",
        latex_text="Структурная формула\nВероятность безотказной работы",
        plain_text="Структурная формула\nВероятность безотказной работы",
        formulas=[
            FormulaItem(
                key="P",
                label="Вероятность безотказной работы",
                symbolic_template="",
                instantiated_latex=r"P_{\text{сист}}(t)=P_{B1}\cdot P_{B2}",
                plain_text="Pсист(t)=P_B1 * P_B2",
            )
        ],
    )
    report = _sample_report()
    report.formula_package = package

    blocks = _formula_render_blocks(report)

    assert len(blocks) == 1
    assert blocks[0]["latex"] == r"P_{\text{сист}}(t)=P_{B1}\cdot P_{B2}"


def test_image_size_normalization_does_not_make_small_images_huge():
    assert _normalize_image_dimensions(100, 40, max_width_px=800, max_height_px=300, max_upscale=1.1) == (110, 44)
    assert _normalize_image_dimensions(1600, 800, max_width_px=800, max_height_px=300, max_upscale=1.1) == (600, 300)


def test_export_xlsx_adds_subscheme_sheet(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    report = _sample_report()
    report.scheme_images = _sample_scheme_images(tmp_path)
    report.scheme_image_path = report.scheme_images[0]["path"]

    target = export_xlsx(tmp_path / "report_subschemes.xlsx", report)
    workbook = openpyxl.load_workbook(target)

    assert "Подсхемы" in workbook.sheetnames
    assert workbook["Подсхемы"]["A1"].value == "Подсхема блока Блок 1"


def test_export_txt_lists_all_scheme_images(tmp_path):
    report = _sample_report()
    report.scheme_images = _sample_scheme_images(tmp_path)
    report.scheme_image_path = report.scheme_images[0]["path"]

    target = export_txt(tmp_path / "report_subschemes.txt", report)
    content = target.read_text(encoding="utf-8")

    assert "Надёжность технических средств" in content
    assert "СХЕМЫ СИСТЕМЫ" in content
    assert "Подсхема блока Блок 1" in content
    assert "Подсхема блока Блок 1.1" in content
